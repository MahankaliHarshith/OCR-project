"""
Excel Report Generator Service.
Generates formatted .xlsx files from receipt data using OpenPyXL.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import (
    EXPORT_DIR,
    EXCEL_HEADER_COLOR,
    EXCEL_HEADER_FONT_COLOR,
    EXCEL_ALT_ROW_COLOR,
    EXCEL_LOW_CONFIDENCE_COLOR,
    EXCEL_MAX_COLUMN_WIDTH,
)

logger = logging.getLogger(__name__)


class ExcelService:
    """
    Generates formatted Excel reports from receipt data.

    Features:
        - Professional header styling
        - Alternating row colors
        - Low-confidence highlighting
        - Summary section with product totals
        - Auto-adjusted column widths
    """

    def __init__(self):
        # Styles
        self.header_fill = PatternFill(
            start_color=EXCEL_HEADER_COLOR,
            end_color=EXCEL_HEADER_COLOR,
            fill_type="solid",
        )
        self.header_font = Font(
            bold=True, color=EXCEL_HEADER_FONT_COLOR, size=12
        )
        self.alt_row_fill = PatternFill(
            start_color=EXCEL_ALT_ROW_COLOR,
            end_color=EXCEL_ALT_ROW_COLOR,
            fill_type="solid",
        )
        self.low_conf_fill = PatternFill(
            start_color=EXCEL_LOW_CONFIDENCE_COLOR,
            end_color=EXCEL_LOW_CONFIDENCE_COLOR,
            fill_type="solid",
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")

    def generate_report(
        self,
        receipts: List[Dict],
        output_path: Optional[str] = None,
    ) -> str:
        """
        Generate a formatted Excel report from receipt data.

        Args:
            receipts: List of receipt dicts. Each must have:
                - receipt_number (str)
                - scan_date (str)
                - scan_time (str, optional)
                - items: list of dicts with code, product, quantity, unit, confidence
            output_path: Custom output file path. Auto-generated if None.

        Returns:
            Path to the generated .xlsx file.
        """
        wb = Workbook()
        logger.debug(
            f"Generating Excel report: {len(receipts)} receipts, "
            f"total_items={sum(len(r.get('items', [])) for r in receipts)}"
        )

        # ── Main data sheet ──────────────────────────────────────────────
        ws_data = wb.active
        ws_data.title = "Daily Sales Report"
        self._build_data_sheet(ws_data, receipts)

        # ── Summary sheet ────────────────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")
        self._build_summary_sheet(ws_summary, receipts)

        # ── Save ─────────────────────────────────────────────────────────
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(EXPORT_DIR / f"Sales_{timestamp}.xlsx")

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Excel report generated: {output_path}")
        return output_path

    # ─── Data Sheet ───────────────────────────────────────────────────────

    def _build_data_sheet(self, ws, receipts: List[Dict]) -> None:
        """Populate the main data sheet with all receipt items."""

        # Title row
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "Daily Sales Report"
        title_cell.font = Font(bold=True, size=16, color=EXCEL_HEADER_COLOR)
        title_cell.alignment = Alignment(horizontal="center")

        # Date row
        ws.merge_cells("A2:H2")
        date_cell = ws["A2"]
        date_cell.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
        date_cell.font = Font(size=10, italic=True, color="666666")
        date_cell.alignment = Alignment(horizontal="center")

        # Header row (row 4)
        headers = [
            "Receipt No",
            "Date",
            "Time",
            "Product Code",
            "Product Name",
            "Quantity",
            "Unit",
            "Confidence",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

        # Data rows (starting row 5)
        row_num = 5
        for receipt in receipts:
            receipt_no = receipt.get("receipt_number", receipt.get("receipt_id", ""))
            date = receipt.get("scan_date", receipt.get("date", ""))
            time_val = receipt.get("scan_time", receipt.get("time", ""))

            for item in receipt.get("items", []):
                values = [
                    receipt_no,
                    date,
                    time_val,
                    item.get("code", ""),
                    item.get("product", ""),
                    item.get("quantity", 0),
                    item.get("unit", "Piece"),
                    f"{item.get('confidence', 0):.1%}",
                ]

                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = self.border

                    # Center-align certain columns
                    if col_num in (1, 4, 6, 7, 8):
                        cell.alignment = self.center_align
                    else:
                        cell.alignment = self.left_align

                # Alternating row colors
                if row_num % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = self.alt_row_fill

                # Highlight low confidence
                confidence = item.get("confidence", 1.0)
                if confidence < 0.85:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = self.low_conf_fill

                row_num += 1

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Freeze header row
        ws.freeze_panes = "A5"

    # ─── Summary Sheet ────────────────────────────────────────────────────

    def _build_summary_sheet(self, ws, receipts: List[Dict]) -> None:
        """Build the summary sheet with product totals."""

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"].value = "Sales Summary"
        ws["A1"].font = Font(bold=True, size=16, color=EXCEL_HEADER_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Stats
        total_receipts = len(receipts)
        total_items = sum(
            len(r.get("items", [])) for r in receipts
        )
        total_quantity = sum(
            item.get("quantity", 0)
            for r in receipts
            for item in r.get("items", [])
        )

        ws.cell(row=3, column=1, value="Total Receipts:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=total_receipts)
        ws.cell(row=4, column=1, value="Total Line Items:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=total_items)
        ws.cell(row=5, column=1, value="Total Quantity:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=total_quantity)

        # Product totals header
        summary_headers = ["Product Code", "Product Name", "Total Quantity", "Unit"]
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

        # Calculate totals by product
        product_totals: Dict[str, Dict] = {}
        for receipt in receipts:
            for item in receipt.get("items", []):
                code = item.get("code", "")
                if code not in product_totals:
                    product_totals[code] = {
                        "name": item.get("product", ""),
                        "quantity": 0,
                        "unit": item.get("unit", "Piece"),
                    }
                product_totals[code]["quantity"] += item.get("quantity", 0)

        # Write product totals
        row = 8
        for code in sorted(product_totals.keys()):
            info = product_totals[code]
            values = [code, info["name"], info["quantity"], info["unit"]]
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_num, value=val)
                cell.border = self.border
                cell.alignment = self.center_align if col_num != 2 else self.left_align

            if row % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.alt_row_fill

            row += 1

        # Grand total row
        ws.cell(row=row + 1, column=1, value="GRAND TOTAL").font = Font(
            bold=True, size=12
        )
        ws.cell(row=row + 1, column=3, value=total_quantity).font = Font(
            bold=True, size=12
        )

        self._adjust_column_widths(ws)

    # ─── Utilities ────────────────────────────────────────────────────────

    def _adjust_column_widths(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass

            adjusted = min(max_length + 3, EXCEL_MAX_COLUMN_WIDTH)
            ws.column_dimensions[column_letter].width = max(adjusted, 10)

    def generate_from_db_receipts(
        self, receipt_ids: List[int], output_path: Optional[str] = None
    ) -> str:
        """
        Generate report from receipt IDs stored in the database.

        Args:
            receipt_ids: List of receipt database IDs.
            output_path: Custom output path.

        Returns:
            Path to generated Excel file.
        """
        from app.database import db

        logger.debug(f"generate_from_db_receipts: receipt_ids={receipt_ids}")
        # Batch fetch all receipts + items in 2 queries (avoids N+1)
        db_receipts = db.get_receipts_batch(receipt_ids)

        receipts = []
        for receipt in db_receipts:
            receipts.append(
                {
                    "receipt_number": receipt["receipt_number"],
                    "scan_date": receipt["scan_date"],
                    "scan_time": receipt["scan_time"],
                    "items": [
                        {
                            "code": item["product_code"],
                            "product": item["product_name"],
                            "quantity": item["quantity"],
                            "unit": item.get("unit", "Piece"),
                            "confidence": item.get("ocr_confidence", 0),
                        }
                        for item in receipt.get("items", [])
                    ],
                }
            )

        if not receipts:
            raise ValueError("No receipts found for the given IDs.")

        return self.generate_report(receipts, output_path)

    def generate_daily_report(
        self, date: Optional[str] = None, output_path: Optional[str] = None
    ) -> str:
        """
        Generate a report for all receipts on a given date.

        Args:
            date: Date string (YYYY-MM-DD). Defaults to today.
            output_path: Custom output path.

        Returns:
            Path to generated Excel file.
        """
        from app.database import db

        if date is None:
            date = datetime.now().strftime("%Y-%m-%d")

        db_receipts = db.get_receipts_by_date(date)

        if not db_receipts:
            raise ValueError(f"No receipts found for date: {date}")

        receipts = []
        for r in db_receipts:
            receipts.append(
                {
                    "receipt_number": r["receipt_number"],
                    "scan_date": r["scan_date"],
                    "scan_time": r["scan_time"],
                    "items": [
                        {
                            "code": item["product_code"],
                            "product": item["product_name"],
                            "quantity": item["quantity"],
                            "unit": item.get("unit", "Piece"),
                            "confidence": item.get("ocr_confidence", 0),
                        }
                        for item in r.get("items", [])
                    ],
                }
            )

        return self.generate_report(receipts, output_path)


# Singleton
excel_service = ExcelService()
